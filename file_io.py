"""File I/O module for reading and exporting data files."""

from io import BytesIO
from typing import BinaryIO, Union
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# Valid file extensions
VALID_EXTENSIONS = {'.xlsx', '.xls', '.csv'}


def get_file_extension(filename: str) -> str:
    """Extract lowercase file extension from filename."""
    if '.' not in filename:
        return ''
    return '.' + filename.rsplit('.', 1)[-1].lower()


def is_valid_file_format(filename: str) -> bool:
    """Check if file has a valid extension (.xlsx, .xls, .csv)."""
    return get_file_extension(filename) in VALID_EXTENSIONS


def read_uploaded_file(file: Union[BinaryIO, BytesIO], filename: str) -> pd.DataFrame:
    """Read an uploaded Excel or CSV file into a DataFrame.
    
    Args:
        file: File-like object containing the data
        filename: Original filename (used to determine format)
        
    Returns:
        DataFrame with the file contents
        
    Raises:
        ValueError: If file format is not supported
    """
    ext = get_file_extension(filename)
    
    if ext not in VALID_EXTENSIONS:
        raise ValueError(f"Unsupported file format: {ext}. Must be .xlsx, .xls, or .csv")
    
    if ext == '.csv':
        return pd.read_csv(file)
    else:
        return pd.read_excel(file)


def read_excel_with_highlights(file: Union[BinaryIO, BytesIO], progress_callback=None) -> tuple[pd.DataFrame, set[tuple[int, int]]]:
    """Read Excel file and detect highlighted cells.
    
    Args:
        file: File-like object containing Excel data
        progress_callback: Optional callback function(percent, message) for progress updates
        
    Returns:
        Tuple of (DataFrame, set of (row_index, col_index) for highlighted cells)
        Row indices are 0-based (matching DataFrame index)
    """
    if progress_callback:
        progress_callback(5, "Reading Excel data...")
    
    # First, read the data quickly with pandas
    df = pd.read_excel(file, engine='openpyxl')
    
    if progress_callback:
        progress_callback(20, "Loading workbook for highlight detection...")
    
    # Reset and load workbook for highlight detection
    file.seek(0)
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    
    highlighted_cells: set[tuple[int, int]] = set()
    
    # Iterate through cells to find highlights (skip header row)
    # Limit to actual data rows for performance
    max_row = min(len(df) + 1, ws.max_row)
    max_col = min(len(df.columns), ws.max_column)
    total_rows = max_row - 1  # Exclude header
    
    if progress_callback:
        progress_callback(30, f"Scanning {total_rows:,} rows for highlights...")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_row + 1, max_col=max_col)):
        # Update progress every 500 rows
        if progress_callback and row_idx > 0 and row_idx % 500 == 0:
            pct = 30 + int((row_idx / total_rows) * 60)  # 30-90% range
            progress_callback(pct, f"Scanning row {row_idx:,} of {total_rows:,}...")
        
        for col_idx, cell in enumerate(row):
            if cell.fill and cell.fill.fill_type == 'solid':
                # Check if fill color is not white/no fill
                fg_color = cell.fill.fgColor
                if fg_color and fg_color.rgb and fg_color.rgb != '00000000' and fg_color.rgb != 'FFFFFFFF':
                    highlighted_cells.add((row_idx, col_idx))
    
    if progress_callback:
        progress_callback(95, "Finalizing...")
    
    wb.close()
    return df, highlighted_cells


def read_excel_fast(file: Union[BinaryIO, BytesIO]) -> pd.DataFrame:
    """Read Excel file quickly without highlight detection.
    
    Args:
        file: File-like object containing Excel data
        
    Returns:
        DataFrame with the file contents
    """
    return pd.read_excel(file, engine='openpyxl')



def export_to_excel(df: pd.DataFrame) -> bytes:
    """Export DataFrame to Excel bytes for download.
    
    Args:
        df: DataFrame to export
        
    Returns:
        Excel file as bytes
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    return output.getvalue()


def export_to_csv(df: pd.DataFrame) -> bytes:
    """Export DataFrame to CSV bytes for download.
    
    Args:
        df: DataFrame to export
        
    Returns:
        CSV file as bytes (UTF-8 encoded)
    """
    return df.to_csv(index=False).encode('utf-8')


# Mapping from internal reason codes to human-readable reasons
REASON_DESCRIPTIONS = {
    'highlighted_cells': 'Highlighted cell in original file',
    'invalid_last_name': 'Invalid last name (must start with letter)',
    'empty_phone': 'Empty/missing phone number',
    'invalid_phone': 'Invalid phone (must be 10 digits, not starting with 1)',
    'invalid_email': 'Invalid email format',
    'dnc_area_code': 'Phone area code matches DNC list',
    'dnc_name_match': 'Name matches DNC list',
    'tcpa_phone_match': 'Phone matches TCPA suppression list',
    'tcpa_zip_match': 'Zip code matches TCPA suppression list',
    'contains_test': 'Contains "TEST" in data',
    'placeholder_email': 'Placeholder email (N/A, No, None, etc.)',
    'prohibited_content': 'Contains prohibited content (loan depot, profanity)',
    'duplicate_phone': 'Duplicate phone number',
    'invalid_uuid': 'Invalid UUID format',
}

# Mapping from reason codes to the column that caused the issue
REASON_TO_COLUMN_FIELD = {
    'invalid_last_name': 'last_name',
    'empty_phone': 'phone',
    'invalid_phone': 'phone',
    'invalid_email': 'email',
    'dnc_area_code': 'phone',
    'dnc_name_match': 'first_name',  # Will highlight both first and last
    'tcpa_phone_match': 'phone',
    'tcpa_zip_match': 'zip_code',
    'placeholder_email': 'email',
    'duplicate_phone': 'phone',
    'invalid_uuid': 'lead_id',
}


def export_removed_rows_to_excel(df: pd.DataFrame, column_mapping) -> bytes:
    """Export removed rows DataFrame to Excel with Reason column and yellow highlighting.
    
    Args:
        df: DataFrame with removed rows (must have '_removal_reason' and '_problem_column' columns)
        column_mapping: ColumnMapping object with mapped column names
        
    Returns:
        Excel file as bytes with Reason column first and problem cells highlighted yellow
    """
    if len(df) == 0:
        return export_to_excel(df)
    
    # Create a copy and prepare the Reason column
    export_df = df.copy()
    
    # Convert internal reason to human-readable
    if '_removal_reason' in export_df.columns:
        export_df['Reason'] = export_df['_removal_reason'].map(
            lambda x: REASON_DESCRIPTIONS.get(x, x)
        )
        # Store original reason for highlighting logic
        original_reasons = export_df['_removal_reason'].tolist()
        export_df = export_df.drop(columns=['_removal_reason'])
    else:
        export_df['Reason'] = 'Unknown'
        original_reasons = ['unknown'] * len(export_df)
    
    # Remove _problem_column if present
    if '_problem_column' in export_df.columns:
        problem_columns = export_df['_problem_column'].tolist()
        export_df = export_df.drop(columns=['_problem_column'])
    else:
        problem_columns = [None] * len(export_df)
    
    # Reorder columns to put Reason first
    cols = export_df.columns.tolist()
    cols.remove('Reason')
    export_df = export_df[['Reason'] + cols]
    
    # Build mapping from field names to actual column names
    field_to_col = {}
    if column_mapping:
        if column_mapping.phone:
            field_to_col['phone'] = column_mapping.phone
        if column_mapping.first_name:
            field_to_col['first_name'] = column_mapping.first_name
        if column_mapping.last_name:
            field_to_col['last_name'] = column_mapping.last_name
        if column_mapping.email:
            field_to_col['email'] = column_mapping.email
        if column_mapping.zip_code:
            field_to_col['zip_code'] = column_mapping.zip_code
        if column_mapping.lead_id:
            field_to_col['lead_id'] = column_mapping.lead_id
    
    # Write to Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='Removed Rows')
        
        # Get the worksheet to apply highlighting
        ws = writer.sheets['Removed Rows']
        
        # Yellow fill for problem cells
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Get column indices in the exported DataFrame
        col_name_to_idx = {col: idx + 1 for idx, col in enumerate(export_df.columns)}  # 1-based for openpyxl
        
        # Apply yellow highlighting to problem cells
        for row_idx, (reason, problem_col) in enumerate(zip(original_reasons, problem_columns)):
            excel_row = row_idx + 2  # +2 because row 1 is header, and enumerate starts at 0
            
            # Determine which column(s) to highlight based on reason
            cols_to_highlight = []
            
            if problem_col and problem_col in col_name_to_idx:
                # Use explicit problem column if provided
                cols_to_highlight.append(problem_col)
            else:
                # Determine from reason code
                field = REASON_TO_COLUMN_FIELD.get(reason)
                if field:
                    actual_col = field_to_col.get(field)
                    if actual_col and actual_col in col_name_to_idx:
                        cols_to_highlight.append(actual_col)
                    
                    # For name match, also highlight last name
                    if reason == 'dnc_name_match':
                        last_col = field_to_col.get('last_name')
                        if last_col and last_col in col_name_to_idx:
                            cols_to_highlight.append(last_col)
            
            # Apply highlighting
            for col_name in cols_to_highlight:
                col_idx = col_name_to_idx[col_name]
                ws.cell(row=excel_row, column=col_idx).fill = yellow_fill
    
    return output.getvalue()
