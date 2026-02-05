"""Property-based tests for refinance data cleansing."""

from hypothesis import given, strategies as st, settings, HealthCheck
import pandas as pd

import sys
sys.path.insert(0, '.')

from file_io import is_valid_file_format, VALID_EXTENSIONS


# **Feature: refinance-data-cleansing, Property 1: File Format Validation**
# **Validates: Requirements 1.2**
# *For any* uploaded file, the system SHALL accept files with extensions 
# .xlsx, .xls, or .csv, and SHALL reject files with any other extension.

# Strategy for generating valid filename bases (alphanumeric)
filename_base_strategy = st.text(
    alphabet=st.characters(whitelist_categories=('L', 'N')),
    min_size=1,
    max_size=20
)


@given(filename_base_strategy)
@settings(max_examples=100)
def test_file_format_validation_accepts_valid_extensions(filename_base: str):
    """Property: Valid extensions (.xlsx, .xls, .csv) are always accepted."""
    for ext in VALID_EXTENSIONS:
        filename = f"{filename_base}{ext}"
        assert is_valid_file_format(filename), f"Should accept {ext} files"


@given(st.text(alphabet='abcdefghijklmnopqrstuvwxyz0123456789', min_size=1, max_size=10))
@settings(max_examples=100)
def test_file_format_validation_rejects_invalid_extensions(ext_without_dot: str):
    """Property: Invalid extensions are always rejected."""
    ext = f".{ext_without_dot}"
    if ext not in VALID_EXTENSIONS:
        filename = f"testfile{ext}"
        assert not is_valid_file_format(filename), f"Should reject {ext} files"


@given(st.text(alphabet='abcdefghijklmnopqrstuvwxyz0123456789_-', min_size=1, max_size=20))
@settings(max_examples=100)
def test_file_format_validation_rejects_no_extension(filename: str):
    """Property: Files without extensions are always rejected."""
    assert not is_valid_file_format(filename), "Should reject files without extension"



from cleaning import (
    normalize_phone, is_valid_last_name, filter_invalid_last_names,
    is_valid_phone, filter_invalid_phones, filter_empty_phones,
    is_valid_email, filter_invalid_emails, remove_highlighted_rows
)


# **Feature: refinance-data-cleansing, Property 16: Phone Normalization Idempotence**
# **Validates: Requirements 3.5**
# *For any* phone string, normalizing it twice SHALL produce the same result as 
# normalizing once, and the result SHALL contain only digits.

@given(st.text(max_size=30))
@settings(max_examples=100)
def test_phone_normalization_idempotence(phone_str: str):
    """Property: Normalizing twice equals normalizing once, result is digits only."""
    once = normalize_phone(phone_str)
    twice = normalize_phone(once)
    
    # Idempotence: f(f(x)) == f(x)
    assert once == twice, f"Normalization not idempotent: '{once}' != '{twice}'"
    
    # Result contains only digits
    assert once.isdigit() or once == '', f"Result contains non-digits: '{once}'"


@given(st.floats(min_value=1000000000, max_value=9999999999, allow_nan=False, allow_infinity=False))
@settings(max_examples=100)
def test_phone_normalization_handles_floats(phone_float: float):
    """Property: Float phone numbers are normalized to digits."""
    result = normalize_phone(phone_float)
    assert result.isdigit() or result == '', f"Float normalization failed: '{result}'"
    # Should be 10 digits for valid phone range
    assert len(result) == 10, f"Expected 10 digits, got {len(result)}"



# **Feature: refinance-data-cleansing, Property 4: Last Name Validation**
# **Validates: Requirements 3.3, 3.4**
# *For any* last name value, the filter SHALL remove the row if and only if:
# (a) the first character is not a letter A-Z or a-z, OR
# (b) the value is empty/whitespace-only, OR
# (c) the value is boolean TRUE.

@given(st.text(alphabet='ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz', min_size=1, max_size=20))
@settings(max_examples=100)
def test_last_name_valid_when_starts_with_letter(name: str):
    """Property: Names starting with letters are valid."""
    assert is_valid_last_name(name), f"Name starting with letter should be valid: '{name}'"


@given(st.text(alphabet='0123456789!@#$%^&*()-_=+[]{}|;:,.<>?/', min_size=1, max_size=10))
@settings(max_examples=100)
def test_last_name_invalid_when_starts_with_non_letter(name: str):
    """Property: Names starting with non-letters are invalid."""
    assert not is_valid_last_name(name), f"Name starting with non-letter should be invalid: '{name}'"


@given(st.text(alphabet=' \t\n\r', min_size=0, max_size=10))
@settings(max_examples=100)
def test_last_name_invalid_when_empty_or_whitespace(name: str):
    """Property: Empty or whitespace-only names are invalid."""
    assert not is_valid_last_name(name), f"Empty/whitespace name should be invalid: '{repr(name)}'"


def test_last_name_invalid_when_boolean_true():
    """Property: Boolean TRUE is invalid."""
    assert not is_valid_last_name(True), "Boolean True should be invalid"
    # String "TRUE" is now valid (starts with letter)
    assert is_valid_last_name("TRUE"), "String 'TRUE' should be valid (starts with letter)"
    assert is_valid_last_name("True"), "String 'True' should be valid"
    assert is_valid_last_name("true"), "String 'true' should be valid"



# **Feature: refinance-data-cleansing, Property 5: Phone Number Validation**
# **Validates: Requirements 3.5, 3.6, 3.7**
# *For any* phone number value, the filter SHALL remove the row if and only if:
# (a) the normalized phone (digits only) is not exactly 10 characters, OR
# (b) the first digit is 1, OR
# (c) the value is empty/missing.

# Strategy for valid 10-digit phones not starting with 1
valid_phone_strategy = st.integers(min_value=2000000000, max_value=9999999999).map(str)

# Strategy for phones starting with 1
phone_starting_with_1_strategy = st.integers(min_value=1000000000, max_value=1999999999).map(str)


@given(valid_phone_strategy)
@settings(max_examples=100)
def test_phone_valid_when_10_digits_not_starting_with_1(phone: str):
    """Property: 10-digit phones not starting with 1 are valid."""
    assert is_valid_phone(phone), f"Valid phone should pass: '{phone}'"


@given(phone_starting_with_1_strategy)
@settings(max_examples=100)
def test_phone_invalid_when_starts_with_1(phone: str):
    """Property: Phones starting with 1 are invalid."""
    assert not is_valid_phone(phone), f"Phone starting with 1 should be invalid: '{phone}'"


@given(st.integers(min_value=100000000, max_value=999999999).map(str))  # 9 digits
@settings(max_examples=100)
def test_phone_invalid_when_not_10_digits_short(phone: str):
    """Property: Phones with less than 10 digits are invalid."""
    assert not is_valid_phone(phone), f"9-digit phone should be invalid: '{phone}'"


@given(st.integers(min_value=10000000000, max_value=99999999999).map(str))  # 11 digits
@settings(max_examples=100)
def test_phone_invalid_when_not_10_digits_long(phone: str):
    """Property: Phones with more than 10 digits are invalid."""
    assert not is_valid_phone(phone), f"11-digit phone should be invalid: '{phone}'"


def test_phone_invalid_when_empty():
    """Property: Empty/None phones are invalid."""
    assert not is_valid_phone(None), "None should be invalid"
    assert not is_valid_phone(""), "Empty string should be invalid"
    assert not is_valid_phone(float('nan')), "NaN should be invalid"



# **Feature: refinance-data-cleansing, Property 6: Email Validation**
# **Validates: Requirements 3.8**
# *For any* email value, the filter SHALL remove the row if and only if the value 
# is empty OR does not contain exactly one @ symbol with characters before and after it.

# Strategy for valid email local parts (before @)
email_local_strategy = st.text(
    alphabet='abcdefghijklmnopqrstuvwxyz0123456789._-',
    min_size=1,
    max_size=20
)

# Strategy for valid email domains (after @)
email_domain_strategy = st.text(
    alphabet='abcdefghijklmnopqrstuvwxyz0123456789.-',
    min_size=3,
    max_size=15
).filter(lambda x: '.' in x and not x.startswith('.') and not x.endswith('.'))


@given(email_local_strategy, st.text(alphabet='abcdefghijklmnopqrstuvwxyz', min_size=2, max_size=10))
@settings(max_examples=100)
def test_email_valid_with_proper_format(local: str, domain: str):
    """Property: Emails with local@domain format are valid."""
    email = f"{local}@{domain}.com"
    assert is_valid_email(email), f"Valid email should pass: '{email}'"


@given(st.text(alphabet='abcdefghijklmnopqrstuvwxyz0123456789', min_size=1, max_size=30).filter(lambda x: '@' not in x))
@settings(max_examples=100)
def test_email_invalid_without_at_symbol(text: str):
    """Property: Text without @ is invalid email."""
    assert not is_valid_email(text), f"Email without @ should be invalid: '{text}'"


@given(st.text(alphabet='abcdefghijklmnopqrstuvwxyz', min_size=1, max_size=10))
@settings(max_examples=100)
def test_email_invalid_with_multiple_at_symbols(local: str):
    """Property: Emails with multiple @ are invalid."""
    email = f"{local}@@domain.com"
    assert not is_valid_email(email), f"Email with multiple @ should be invalid: '{email}'"


def test_email_invalid_when_empty():
    """Property: Empty/None emails are invalid."""
    assert not is_valid_email(None), "None should be invalid"
    assert not is_valid_email(""), "Empty string should be invalid"
    assert not is_valid_email("   "), "Whitespace should be invalid"


def test_email_invalid_with_nothing_before_or_after_at():
    """Property: Emails with nothing before or after @ are invalid."""
    assert not is_valid_email("@domain.com"), "Nothing before @ should be invalid"
    assert not is_valid_email("user@"), "Nothing after @ should be invalid"
    assert not is_valid_email("@"), "Just @ should be invalid"



# **Feature: refinance-data-cleansing, Property 3: Highlighted Row Removal**
# **Validates: Requirements 3.2**
# *For any* DataFrame with highlighted cells, after filtering, no row that 
# contained a highlighted cell SHALL remain in the output.

@given(
    st.lists(st.tuples(st.integers(0, 9), st.integers(0, 2)), min_size=0, max_size=10, unique=True)
)
@settings(max_examples=100)
def test_highlighted_rows_removed(highlighted_coords: list):
    """Property: All rows with highlighted cells are removed."""
    # Create a test DataFrame with 10 rows, 3 columns
    df = pd.DataFrame({
        'A': range(10),
        'B': range(10, 20),
        'C': range(20, 30)
    })
    
    highlighted_cells = set(highlighted_coords)
    result = remove_highlighted_rows(df, highlighted_cells)
    
    # Get rows that should be removed
    highlighted_row_indices = {row_idx for row_idx, _ in highlighted_cells}
    
    # Verify no highlighted rows remain
    for row_idx in highlighted_row_indices:
        assert row_idx not in result.cleaned_df.index, f"Row {row_idx} should have been removed"
    
    # Verify removed count matches
    assert result.removed_count == len(highlighted_row_indices)
    
    # Verify total rows preserved
    assert len(result.cleaned_df) + len(result.removed_df) == len(df)


def test_highlighted_rows_empty_set():
    """Property: Empty highlight set removes no rows."""
    df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
    result = remove_highlighted_rows(df, set())
    
    assert len(result.cleaned_df) == 3
    assert result.removed_count == 0



from matching import (
    filter_by_area_code, filter_by_name_match, filter_by_tcpa_phones,
    filter_by_tcpa_zips, normalize_name, normalize_zip, load_phones_from_all_tabs
)
from io import BytesIO
import openpyxl


# **Feature: refinance-data-cleansing, Property 7: Area Code Matching**
# **Validates: Requirements 4.1**
# *For any* phone number and DNC area code set, the filter SHALL remove the row 
# if and only if the first 3 digits of the normalized phone exist in the area code set.

@given(
    st.lists(st.integers(min_value=2000000000, max_value=9999999999).map(str), min_size=1, max_size=10),
    st.lists(st.text(alphabet='0123456789', min_size=3, max_size=3), min_size=1, max_size=5, unique=True)
)
@settings(max_examples=100)
def test_area_code_matching(phones: list, area_codes: list):
    """Property: Rows are removed iff phone area code is in DNC set."""
    df = pd.DataFrame({'Phone': phones})
    dnc_set = set(area_codes)
    
    result = filter_by_area_code(df, 'Phone', dnc_set)
    
    # Verify: all removed rows have matching area codes
    for _, row in result.removed_df.iterrows():
        phone = normalize_phone(row['Phone'])
        assert phone[:3] in dnc_set, f"Removed phone {phone} doesn't match any area code"
    
    # Verify: all kept rows don't have matching area codes
    for _, row in result.cleaned_df.iterrows():
        phone = normalize_phone(row['Phone'])
        assert phone[:3] not in dnc_set, f"Kept phone {phone} matches area code but wasn't removed"



# **Feature: refinance-data-cleansing, Property 8: Name Matching**
# **Validates: Requirements 4.2**
# *For any* first name, last name, and DNC name set, the filter SHALL remove the row 
# if and only if the normalized concatenation (lowercase, trimmed) of first+last name exists in the name set.

@given(
    st.lists(st.tuples(
        st.text(alphabet='ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz', min_size=1, max_size=10),
        st.text(alphabet='ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz', min_size=1, max_size=10)
    ), min_size=1, max_size=10)
)
@settings(max_examples=100)
def test_name_matching(name_pairs: list):
    """Property: Rows are removed iff concatenated name is in DNC set."""
    first_names = [p[0] for p in name_pairs]
    last_names = [p[1] for p in name_pairs]
    df = pd.DataFrame({'FirstName': first_names, 'LastName': last_names})
    
    # Create DNC set from some of the names (first half)
    dnc_names = set()
    for i in range(len(name_pairs) // 2):
        concat = normalize_name(first_names[i]) + normalize_name(last_names[i])
        dnc_names.add(concat)
    
    result = filter_by_name_match(df, 'FirstName', 'LastName', dnc_names)
    
    # Verify: all removed rows have matching names
    for _, row in result.removed_df.iterrows():
        concat = normalize_name(row['FirstName']) + normalize_name(row['LastName'])
        assert concat in dnc_names, f"Removed name {concat} not in DNC set"
    
    # Verify: all kept rows don't have matching names
    for _, row in result.cleaned_df.iterrows():
        concat = normalize_name(row['FirstName']) + normalize_name(row['LastName'])
        assert concat not in dnc_names, f"Kept name {concat} is in DNC set but wasn't removed"



# **Feature: refinance-data-cleansing, Property 9: TCPA Phone Matching**
# **Validates: Requirements 4.3**
# *For any* phone number and TCPA phone set, the filter SHALL remove the row 
# if and only if the normalized phone exists in the TCPA phone set.

@given(
    st.lists(st.integers(min_value=2000000000, max_value=9999999999).map(str), min_size=1, max_size=10)
)
@settings(max_examples=100)
def test_tcpa_phone_matching(phones: list):
    """Property: Rows are removed iff phone is in TCPA set."""
    df = pd.DataFrame({'Phone': phones})
    
    # Create TCPA set from some phones (first half)
    tcpa_phones = set(normalize_phone(p) for p in phones[:len(phones)//2])
    
    result = filter_by_tcpa_phones(df, 'Phone', tcpa_phones)
    
    # Verify: all removed rows have matching phones
    for _, row in result.removed_df.iterrows():
        phone = normalize_phone(row['Phone'])
        assert phone in tcpa_phones, f"Removed phone {phone} not in TCPA set"
    
    # Verify: all kept rows don't have matching phones
    for _, row in result.cleaned_df.iterrows():
        phone = normalize_phone(row['Phone'])
        assert phone not in tcpa_phones, f"Kept phone {phone} is in TCPA set but wasn't removed"



# **Feature: refinance-data-cleansing, Property 10: TCPA Zip Matching**
# **Validates: Requirements 4.4**
# *For any* zip code and TCPA zip set, the filter SHALL remove the row 
# if and only if the normalized zip (first 5 digits) exists in the TCPA zip set.

@given(
    st.lists(st.integers(min_value=10000, max_value=99999).map(str), min_size=1, max_size=10)
)
@settings(max_examples=100)
def test_tcpa_zip_matching(zips: list):
    """Property: Rows are removed iff zip is in TCPA set."""
    df = pd.DataFrame({'ZipCode': zips})
    
    # Create TCPA set from some zips (first half)
    tcpa_zips = set(normalize_zip(z) for z in zips[:len(zips)//2])
    
    result = filter_by_tcpa_zips(df, 'ZipCode', tcpa_zips)
    
    # Verify: all removed rows have matching zips
    for _, row in result.removed_df.iterrows():
        zip_code = normalize_zip(row['ZipCode'])
        assert zip_code in tcpa_zips, f"Removed zip {zip_code} not in TCPA set"
    
    # Verify: all kept rows don't have matching zips
    for _, row in result.cleaned_df.iterrows():
        zip_code = normalize_zip(row['ZipCode'])
        assert zip_code not in tcpa_zips, f"Kept zip {zip_code} is in TCPA set but wasn't removed"



from cleaning import (
    filter_test_entries, filter_placeholder_emails, filter_prohibited_content,
    remove_duplicate_phones, filter_invalid_uuid, is_valid_uuid, PLACEHOLDER_EMAILS
)


# **Feature: refinance-data-cleansing, Property 11: TEST Entry Detection**
# **Validates: Requirements 4.5**
# *For any* first or last name value, the filter SHALL remove the row if and only if 
# the lowercase value contains the substring "test".

@given(st.lists(st.text(alphabet='abcdefghijklmnopqrstuvwxyz', min_size=1, max_size=20), min_size=1, max_size=10))
@settings(max_examples=100)
def test_test_entry_detection_in_names(values: list):
    """Property: Rows containing 'test' in first/last name are removed."""
    df = pd.DataFrame({'FirstName': values, 'LastName': ['Smith'] * len(values)})
    result = filter_test_entries(df, 'FirstName', 'LastName')
    
    # Verify: all removed rows contain 'test' in first name
    for _, row in result.removed_df.iterrows():
        assert 'test' in str(row['FirstName']).lower(), f"Removed row doesn't contain 'test': {row['FirstName']}"
    
    # Verify: all kept rows don't contain 'test' in first name
    for _, row in result.cleaned_df.iterrows():
        assert 'test' not in str(row['FirstName']).lower(), f"Kept row contains 'test': {row['FirstName']}"


def test_test_entry_case_insensitive():
    """Property: TEST detection is case-insensitive."""
    df = pd.DataFrame({
        'FirstName': ['TEST', 'Test', 'test', 'TeSt', 'normal'],
        'LastName': ['Smith', 'Jones', 'Brown', 'Davis', 'Wilson']
    })
    result = filter_test_entries(df, 'FirstName', 'LastName')
    
    assert result.removed_count == 4
    assert len(result.cleaned_df) == 1


def test_test_entry_in_last_name():
    """Property: TEST in last name is also detected."""
    df = pd.DataFrame({
        'FirstName': ['John', 'Jane'],
        'LastName': ['TestUser', 'Smith']
    })
    result = filter_test_entries(df, 'FirstName', 'LastName')
    
    assert result.removed_count == 1
    assert len(result.cleaned_df) == 1


def test_test_entry_only_checks_name_columns():
    """Property: TEST in other columns is NOT detected."""
    df = pd.DataFrame({
        'FirstName': ['John', 'Jane'],
        'LastName': ['Smith', 'Doe'],
        'Notes': ['This is a test', 'Normal notes']
    })
    result = filter_test_entries(df, 'FirstName', 'LastName')
    
    # Should NOT remove any rows since TEST is only in Notes column
    assert result.removed_count == 0
    assert len(result.cleaned_df) == 2



# **Feature: refinance-data-cleansing, Property 12: Placeholder Email Detection**
# **Validates: Requirements 4.6**
# *For any* email value, the filter SHALL remove the row if and only if the normalized 
# value (lowercase, trimmed) equals one of: "n/a", "no", "nada", "na", "noemail", "none".

@given(st.sampled_from(list(PLACEHOLDER_EMAILS)))
@settings(max_examples=100)
def test_placeholder_email_detected(placeholder: str):
    """Property: Placeholder emails are detected and removed."""
    # Test various case variations
    variations = [placeholder, placeholder.upper(), placeholder.title(), f"  {placeholder}  "]
    df = pd.DataFrame({'Email': variations})
    result = filter_placeholder_emails(df, 'Email')
    
    assert result.removed_count == len(variations), f"All variations of '{placeholder}' should be removed"


@given(st.emails())
@settings(max_examples=100)
def test_valid_email_not_placeholder(email: str):
    """Property: Valid emails are not detected as placeholders."""
    df = pd.DataFrame({'Email': [email]})
    result = filter_placeholder_emails(df, 'Email')
    
    # Valid emails should not be removed (unless they happen to match a placeholder)
    normalized = email.strip().lower()
    if normalized not in PLACEHOLDER_EMAILS:
        assert result.removed_count == 0, f"Valid email '{email}' should not be removed"



# **Feature: refinance-data-cleansing, Property 13: Prohibited Content Detection**
# **Validates: Requirements 4.7**
# *For any* text field value, the filter SHALL remove the row if and only if 
# the lowercase value contains "loan depot" or "fuck".

def test_prohibited_content_loan_depot():
    """Property: 'loan depot' content is detected."""
    df = pd.DataFrame({'Field': ['loan depot', 'LOAN DEPOT', 'Loan Depot', 'contains loan depot here', 'normal']})
    result = filter_prohibited_content(df)
    
    assert result.removed_count == 4
    assert len(result.cleaned_df) == 1


def test_prohibited_content_profanity():
    """Property: Profanity is detected."""
    df = pd.DataFrame({'Field': ['fuck', 'FUCK', 'what the fuck', 'normal', 'clean']})
    result = filter_prohibited_content(df)
    
    assert result.removed_count == 3
    assert len(result.cleaned_df) == 2


@given(st.text(alphabet='abcdefghijklmnopqrstuvwxyz ', min_size=1, max_size=30).filter(
    lambda x: 'loan depot' not in x.lower() and 'fuck' not in x.lower()
))
@settings(max_examples=100)
def test_clean_content_not_removed(text: str):
    """Property: Clean content is not removed."""
    df = pd.DataFrame({'Field': [text]})
    result = filter_prohibited_content(df)
    
    assert result.removed_count == 0, f"Clean text '{text}' should not be removed"



# **Feature: refinance-data-cleansing, Property 14: Duplicate Phone Removal**
# **Validates: Requirements 4.8**
# *For any* DataFrame with N rows containing D unique phone numbers, after duplicate 
# removal the output SHALL have exactly D rows, with each unique phone appearing exactly once.

@given(st.lists(st.integers(min_value=2000000000, max_value=9999999999).map(str), min_size=1, max_size=20))
@settings(max_examples=100)
def test_duplicate_phone_removal(phones: list):
    """Property: After dedup, unique phones = unique rows."""
    df = pd.DataFrame({'Phone': phones, 'Data': range(len(phones))})
    
    unique_phones = set(normalize_phone(p) for p in phones)
    result = remove_duplicate_phones(df, 'Phone')
    
    # Verify: output has exactly as many rows as unique phones
    assert len(result.cleaned_df) == len(unique_phones), \
        f"Expected {len(unique_phones)} rows, got {len(result.cleaned_df)}"
    
    # Verify: each unique phone appears exactly once
    result_phones = [normalize_phone(p) for p in result.cleaned_df['Phone']]
    assert len(result_phones) == len(set(result_phones)), "Duplicate phones remain in output"
    
    # Verify: total rows preserved
    assert len(result.cleaned_df) + len(result.removed_df) == len(df)



# **Feature: refinance-data-cleansing, Property 15: UUID Format Validation**
# **Validates: Requirements 4.9**
# *For any* Universal Lead ID value, the filter SHALL remove the row if and only if 
# the value does not match the regex pattern ^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$.

@given(st.uuids())
@settings(max_examples=100)
def test_valid_uuid_accepted(uuid_val):
    """Property: Valid UUIDs are accepted."""
    uuid_str = str(uuid_val).upper()  # Test uppercase
    assert is_valid_uuid(uuid_str), f"Valid UUID should be accepted: {uuid_str}"
    
    uuid_str_lower = str(uuid_val).lower()  # Test lowercase
    assert is_valid_uuid(uuid_str_lower), f"Valid UUID should be accepted: {uuid_str_lower}"


@given(st.text(alphabet='0123456789abcdef-', min_size=1, max_size=40).filter(
    lambda x: len(x) != 36 or x.count('-') != 4
))
@settings(max_examples=100)
def test_invalid_uuid_rejected(text: str):
    """Property: Invalid UUIDs are rejected."""
    # Skip if it accidentally matches UUID format
    if len(text) == 36 and text.count('-') == 4:
        return
    assert not is_valid_uuid(text), f"Invalid UUID should be rejected: {text}"


def test_uuid_filter_removes_invalid():
    """Property: Filter removes rows with invalid UUIDs."""
    df = pd.DataFrame({
        'UUID': [
            'F88CC4BA-95B2-353F-9AE2-7894C12BDCCD',  # Valid
            '53CB2C6F-DC72-1C1F-D5A8-DA507CF67099',  # Valid
            'invalid-uuid',  # Invalid
            '12345',  # Invalid
            None,  # Invalid
        ]
    })
    result = filter_invalid_uuid(df, 'UUID')
    
    assert len(result.cleaned_df) == 2
    assert result.removed_count == 3



# **Feature: multi-file-workflow, Unit Tests for load_phones_from_all_tabs()**
# **Validates: Requirements 5.2, 5.3**

def create_excel_with_tabs(tabs_data: dict) -> BytesIO:
    """Helper to create an Excel file with multiple tabs.
    
    Args:
        tabs_data: Dict mapping sheet_name -> list of phone values
        
    Returns:
        BytesIO containing the Excel file
    """
    output = BytesIO()
    workbook = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    for sheet_name, phones in tabs_data.items():
        ws = workbook.create_sheet(title=sheet_name)
        ws.cell(row=1, column=1, value='Phone')
        for i, phone in enumerate(phones, start=2):
            ws.cell(row=i, column=1, value=phone)
    
    workbook.save(output)
    output.seek(0)
    return output


def test_load_phones_from_single_tab():
    """Test loading phones from a single tab."""
    excel_file = create_excel_with_tabs({
        'Sheet1': ['5551234567', '5559876543', '5551112222']
    })
    
    result = load_phones_from_all_tabs(excel_file)
    
    assert len(result) == 3
    assert '5551234567' in result
    assert '5559876543' in result
    assert '5551112222' in result


def test_load_phones_from_multiple_tabs():
    """Test loading phones from multiple tabs - all tabs should be read."""
    excel_file = create_excel_with_tabs({
        'Tab1': ['5551111111', '5552222222'],
        'Tab2': ['5553333333', '5554444444'],
        'Tab3': ['5555555555']
    })
    
    result = load_phones_from_all_tabs(excel_file)
    
    # Should have all 5 phones from all 3 tabs
    assert len(result) == 5
    assert '5551111111' in result
    assert '5552222222' in result
    assert '5553333333' in result
    assert '5554444444' in result
    assert '5555555555' in result


def test_load_phones_normalizes_formats():
    """Test that phone numbers are normalized to 10 digits."""
    excel_file = create_excel_with_tabs({
        'Sheet1': [
            '(555) 123-4567',      # Formatted with parens and dashes
            '555-987-6543',        # Formatted with dashes
            '555.111.2222',        # Formatted with dots
            5553334444,            # Integer
            5.556667777e9,         # Scientific notation (float)
        ]
    })
    
    result = load_phones_from_all_tabs(excel_file)
    
    # All should be normalized to 10 digits
    assert '5551234567' in result
    assert '5559876543' in result
    assert '5551112222' in result
    assert '5553334444' in result
    assert '5556667777' in result


def test_load_phones_excludes_invalid():
    """Test that invalid phone numbers are excluded."""
    excel_file = create_excel_with_tabs({
        'Sheet1': [
            '5551234567',    # Valid 10 digits
            '123456789',     # Only 9 digits - invalid
            '12345678901',   # 11 digits - invalid
            '',              # Empty - invalid
            None,            # None - invalid
            'not a phone',   # Text - invalid
        ]
    })
    
    result = load_phones_from_all_tabs(excel_file)
    
    # Only the valid 10-digit phone should be included
    assert len(result) == 1
    assert '5551234567' in result


def test_load_phones_deduplicates_across_tabs():
    """Test that duplicate phones across tabs are deduplicated."""
    excel_file = create_excel_with_tabs({
        'Tab1': ['5551234567', '5559876543'],
        'Tab2': ['5551234567', '5551112222'],  # 5551234567 is duplicate
        'Tab3': ['5559876543', '5553334444'],  # 5559876543 is duplicate
    })
    
    result = load_phones_from_all_tabs(excel_file)
    
    # Should have 4 unique phones (not 6)
    assert len(result) == 4
    assert '5551234567' in result
    assert '5559876543' in result
    assert '5551112222' in result
    assert '5553334444' in result


def test_load_phones_handles_empty_tabs():
    """Test that empty tabs are handled gracefully."""
    excel_file = create_excel_with_tabs({
        'Tab1': ['5551234567'],
        'EmptyTab': [],
        'Tab3': ['5559876543'],
    })
    
    result = load_phones_from_all_tabs(excel_file)
    
    # Should have phones from non-empty tabs
    assert len(result) == 2
    assert '5551234567' in result
    assert '5559876543' in result


def test_load_phones_finds_phone_column_by_name():
    """Test that the function finds columns with 'phone' in the name."""
    output = BytesIO()
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = 'Sheet1'
    
    # Create columns with different names
    ws.cell(row=1, column=1, value='ID')
    ws.cell(row=1, column=2, value='PhoneNumber')  # Should be detected
    ws.cell(row=1, column=3, value='Name')
    
    ws.cell(row=2, column=1, value='1')
    ws.cell(row=2, column=2, value='5551234567')
    ws.cell(row=2, column=3, value='John')
    
    ws.cell(row=3, column=1, value='2')
    ws.cell(row=3, column=2, value='5559876543')
    ws.cell(row=3, column=3, value='Jane')
    
    workbook.save(output)
    output.seek(0)
    
    result = load_phones_from_all_tabs(output)
    
    assert len(result) == 2
    assert '5551234567' in result
    assert '5559876543' in result


def test_load_phones_uses_first_column_as_fallback():
    """Test that first column is used if no 'phone' column exists."""
    output = BytesIO()
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = 'Sheet1'
    
    # Create columns without 'phone' in name
    ws.cell(row=1, column=1, value='Numbers')  # First column - should be used
    ws.cell(row=1, column=2, value='Name')
    
    ws.cell(row=2, column=1, value='5551234567')
    ws.cell(row=2, column=2, value='John')
    
    ws.cell(row=3, column=1, value='5559876543')
    ws.cell(row=3, column=2, value='Jane')
    
    workbook.save(output)
    output.seek(0)
    
    result = load_phones_from_all_tabs(output)
    
    assert len(result) == 2
    assert '5551234567' in result
    assert '5559876543' in result
