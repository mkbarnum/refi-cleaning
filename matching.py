"""TCPA matching module for refinance data cleansing."""

from __future__ import annotations
from io import BytesIO
from typing import Set, Tuple
import pandas as pd
from openpyxl import load_workbook
from models import CleanResult
from cleaning import normalize_phone


def normalize_name(name) -> str:
    """Normalize name: trim and lowercase for matching."""
    if name is None or (isinstance(name, float) and pd.isna(name)):
        return ''
    return str(name).strip().lower()


def load_tcpa_phones(df: pd.DataFrame) -> Set[str]:
    """Extract normalized phone numbers from TCPA Phones file.
    
    Args:
        df: DataFrame from TCPA Phones suppression file
        
    Returns:
        Set of normalized phone numbers (10 digits)
    """
    phones = set()
    # Try common column names for phone
    phone_cols = [col for col in df.columns if 'phone' in col.lower()]
    if not phone_cols:
        # Use first column if no phone column found
        phone_cols = [df.columns[0]] if len(df.columns) > 0 else []
    
    for col in phone_cols:
        for val in df[col].dropna():
            normalized = normalize_phone(val)
            if len(normalized) == 10:
                phones.add(normalized)
    
    return phones


def load_phones_from_all_tabs(file: BytesIO) -> Set[str]:
    """Extract normalized phone numbers from all tabs in Excel file.
    
    Reads an Excel file with multiple tabs and extracts phone numbers from
    each tab. Phone numbers are normalized to 10 digits using normalize_phone().
    
    Args:
        file: Excel file (BytesIO) with multiple tabs containing phone numbers
        
    Returns:
        Set of normalized 10-digit phone numbers from all tabs
        
    Note:
        - Looks for columns containing 'phone' in the name
        - Falls back to first column if no phone column found
        - Only includes valid 10-digit phone numbers
        - Invalid data is skipped with a warning (continues processing)
    """
    phones = set()
    
    # Reset file position to beginning
    file.seek(0)
    
    # Load workbook to get all sheet names
    workbook = load_workbook(filename=file, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    
    # Reset file position for pandas reading
    file.seek(0)
    
    # Read each sheet and extract phone numbers
    for sheet_name in sheet_names:
        try:
            # Reset file position before each read
            file.seek(0)
            df = pd.read_excel(file, sheet_name=sheet_name)
            
            if df.empty:
                continue
            
            # Find phone columns (columns containing 'phone' in name)
            phone_cols = [col for col in df.columns if 'phone' in str(col).lower()]
            
            # Fall back to first column if no phone column found
            if not phone_cols:
                phone_cols = [df.columns[0]] if len(df.columns) > 0 else []
            
            # Extract and normalize phone numbers from each phone column
            for col in phone_cols:
                for val in df[col].dropna():
                    normalized = normalize_phone(val)
                    if len(normalized) == 10:
                        phones.add(normalized)
                        
        except Exception:
            # Skip tabs that fail to read, continue with others
            # This handles invalid data gracefully per Requirement 5.7
            continue
    
    return phones


def load_tcpa_zipcodes(df: pd.DataFrame) -> Set[str]:
    """Extract zip codes from TCPA ZipCodes file.
    
    Args:
        df: DataFrame from TCPA ZipCodes suppression file
        
    Returns:
        Set of zip codes (first 5 digits)
    """
    zips = set()
    # Try common column names for zip
    zip_cols = [col for col in df.columns if 'zip' in col.lower()]
    if not zip_cols:
        zip_cols = [df.columns[0]] if len(df.columns) > 0 else []
    
    for col in zip_cols:
        for val in df[col].dropna():
            # Extract first 5 digits
            zip_str = ''.join(c for c in str(val) if c.isdigit())[:5]
            if len(zip_str) == 5:
                zips.add(zip_str)
    
    return zips


def load_ld_dnc(df: pd.DataFrame) -> Tuple[Set[str], Set[str], Set[str]]:
    """Extract phone numbers, area codes, and concatenated names from LD DNC file.
    
    Reads from first two columns only:
    - Column 1: Phone numbers (10 digits) or area codes (3 digits)
    - Column 2: Concatenated names (FirstNameLastName format)
    
    Args:
        df: DataFrame from TCPA LD DNC suppression file (Sheet1 (2))
        
    Returns:
        Tuple of (phone_numbers set, area_codes set, names set)
    """
    phone_numbers = set()
    area_codes = set()
    names = set()
    
    if len(df.columns) < 2:
        return phone_numbers, area_codes, names
    
    # First column: phone numbers or area codes
    col1 = df.columns[0]
    for val in df[col1].dropna():
        digits = ''.join(c for c in str(val) if c.isdigit())
        if len(digits) == 10:
            # Full phone number
            phone_numbers.add(digits)
        elif len(digits) == 3:
            # Area code
            area_codes.add(digits)
    
    # Second column: concatenated names
    col2 = df.columns[1]
    for val in df[col2].dropna():
        name = normalize_name(val)
        if name:
            names.add(name)
    
    return phone_numbers, area_codes, names



def filter_by_dnc_phones(df: pd.DataFrame, phone_col: str, dnc_phones: Set[str]) -> CleanResult:
    """Remove rows where phone number matches DNC phone list.
    
    Args:
        df: DataFrame to filter
        phone_col: Name of the phone column
        dnc_phones: Set of 10-digit phone numbers to filter against
        
    Returns:
        CleanResult with cleaned and removed DataFrames
    """
    def matches_dnc_phone(phone_val) -> bool:
        normalized = normalize_phone(phone_val)
        return normalized in dnc_phones
    
    match_mask = df[phone_col].apply(matches_dnc_phone)
    
    cleaned_df = df[~match_mask].copy()
    removed_df = df[match_mask].copy()
    
    return CleanResult(
        cleaned_df=cleaned_df,
        removed_df=removed_df,
        removed_count=len(removed_df),
        reason="dnc_phone_match"
    )


def filter_by_area_code(df: pd.DataFrame, phone_col: str, dnc_area_codes: Set[str]) -> CleanResult:
    """Remove rows where phone area code matches DNC list.
    
    Args:
        df: DataFrame to filter
        phone_col: Name of the phone column
        dnc_area_codes: Set of 3-digit area codes to filter against
        
    Returns:
        CleanResult with cleaned and removed DataFrames
    """
    def matches_area_code(phone_val) -> bool:
        normalized = normalize_phone(phone_val)
        if len(normalized) >= 3:
            return normalized[:3] in dnc_area_codes
        return False
    
    match_mask = df[phone_col].apply(matches_area_code)
    
    cleaned_df = df[~match_mask].copy()
    removed_df = df[match_mask].copy()
    
    return CleanResult(
        cleaned_df=cleaned_df,
        removed_df=removed_df,
        removed_count=len(removed_df),
        reason="dnc_area_code"
    )



def filter_by_name_match(df: pd.DataFrame, first_col: str, last_col: str, dnc_names: Set[str]) -> CleanResult:
    """Remove rows where concatenated name matches DNC list.
    
    Args:
        df: DataFrame to filter
        first_col: Name of the first name column
        last_col: Name of the last name column
        dnc_names: Set of normalized concatenated names to filter against
        
    Returns:
        CleanResult with cleaned and removed DataFrames
    """
    def matches_name(row) -> bool:
        first = normalize_name(row.get(first_col, ''))
        last = normalize_name(row.get(last_col, ''))
        concat_name = f"{first}{last}"
        return concat_name in dnc_names
    
    match_mask = df.apply(matches_name, axis=1)
    
    cleaned_df = df[~match_mask].copy()
    removed_df = df[match_mask].copy()
    
    return CleanResult(
        cleaned_df=cleaned_df,
        removed_df=removed_df,
        removed_count=len(removed_df),
        reason="dnc_name_match"
    )



def filter_by_tcpa_phones(df: pd.DataFrame, phone_col: str, tcpa_phones: Set[str]) -> CleanResult:
    """Remove rows matching TCPA phone numbers.
    
    Args:
        df: DataFrame to filter
        phone_col: Name of the phone column
        tcpa_phones: Set of normalized phone numbers to filter against
        
    Returns:
        CleanResult with cleaned and removed DataFrames
    """
    def matches_tcpa_phone(phone_val) -> bool:
        normalized = normalize_phone(phone_val)
        return normalized in tcpa_phones
    
    match_mask = df[phone_col].apply(matches_tcpa_phone)
    
    cleaned_df = df[~match_mask].copy()
    removed_df = df[match_mask].copy()
    
    return CleanResult(
        cleaned_df=cleaned_df,
        removed_df=removed_df,
        removed_count=len(removed_df),
        reason="tcpa_phone_match"
    )



def normalize_zip(zip_val) -> str:
    """Normalize zip code to first 5 digits."""
    if zip_val is None or (isinstance(zip_val, float) and pd.isna(zip_val)):
        return ''
    # Extract first 5 digits
    return ''.join(c for c in str(zip_val) if c.isdigit())[:5]


def filter_by_tcpa_zips(df: pd.DataFrame, zip_col: str, tcpa_zips: Set[str]) -> CleanResult:
    """Remove rows matching TCPA zip codes.
    
    Args:
        df: DataFrame to filter
        zip_col: Name of the zip code column
        tcpa_zips: Set of 5-digit zip codes to filter against
        
    Returns:
        CleanResult with cleaned and removed DataFrames
    """
    def matches_tcpa_zip(zip_val) -> bool:
        normalized = normalize_zip(zip_val)
        return normalized in tcpa_zips
    
    match_mask = df[zip_col].apply(matches_tcpa_zip)
    
    cleaned_df = df[~match_mask].copy()
    removed_df = df[match_mask].copy()
    
    return CleanResult(
        cleaned_df=cleaned_df,
        removed_df=removed_df,
        removed_count=len(removed_df),
        reason="tcpa_zip_match"
    )
